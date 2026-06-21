"""
文件处理插件
支持上传、解析、处理各种文件格式
"""

import os
import json
from typing import Dict, Any, List, Optional
from pathlib import Path


class FileHandlerPlugin:
    """文件处理插件"""
    
    SUPPORTED_FORMATS = {
        "text": [".txt", ".md", ".csv", ".json", ".xml", ".yaml", ".yml"],
        "document": [".pdf", ".doc", ".docx", ".ppt", ".pptx", ".xls", ".xlsx"],
        "image": [".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"],
        "code": [".py", ".js", ".ts", ".java", ".cpp", ".c", ".go", ".rs", ".html", ".css"],
    }
    
    def __init__(self):
        self.upload_dir = os.path.expanduser("~/.iqra/uploads")
        os.makedirs(self.upload_dir, exist_ok=True)
        self.file_cache: Dict[str, Dict[str, Any]] = {}
    
    def get_file_type(self, filepath: str) -> str:
        """获取文件类型"""
        ext = Path(filepath).suffix.lower()
        for ftype, exts in self.SUPPORTED_FORMATS.items():
            if ext in exts:
                return ftype
        return "unknown"
    
    def upload_file(self, filepath: str, session_id: str = "default") -> Dict[str, Any]:
        """上传文件到会话"""
        if not os.path.exists(filepath):
            return {"success": False, "error": "文件不存在"}
        
        file_id = f"{session_id}_{os.path.basename(filepath)}"
        file_type = self.get_file_type(filepath)
        
        # 复制到上传目录
        import shutil
        dest = os.path.join(self.upload_dir, file_id)
        shutil.copy2(filepath, dest)
        
        info = {
            "id": file_id,
            "name": os.path.basename(filepath),
            "path": dest,
            "type": file_type,
            "size": os.path.getsize(filepath),
            "session": session_id,
        }
        self.file_cache[file_id] = info
        
        return {"success": True, "file_info": info}
    
    def read_file(self, file_id: str) -> Dict[str, Any]:
        """读取文件内容"""
        if file_id not in self.file_cache:
            return {"success": False, "error": "文件未找到"}
        
        info = self.file_cache[file_id]
        filepath = info["path"]
        file_type = info["type"]
        
        try:
            if file_type == "text" or file_type == "code":
                with open(filepath, "r", encoding="utf-8") as f:
                    content = f.read()
                return {"success": True, "content": content, "type": file_type}
            
            elif file_type == "image":
                # 返回图片路径，供显示使用
                return {"success": True, "content": filepath, "type": "image"}
            
            elif file_type == "document":
                # 尝试提取文本
                return self._extract_document_text(filepath)
            
            else:
                return {"success": False, "error": f"不支持的文件类型: {file_type}"}
                
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _extract_document_text(self, filepath: str) -> Dict[str, Any]:
        """提取文档文本"""
        ext = Path(filepath).suffix.lower()
        
        try:
            if ext == ".pdf":
                return self._read_pdf(filepath)
            elif ext in [".docx", ".doc"]:
                return self._read_word(filepath)
            elif ext in [".xlsx", ".xls"]:
                return self._read_excel(filepath)
            else:
                return {"success": False, "error": f"暂不支持提取 {ext} 格式"}
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _read_pdf(self, filepath: str) -> Dict[str, Any]:
        """读取 PDF"""
        try:
            import PyPDF2
            with open(filepath, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() + "\n"
                return {"success": True, "content": text, "type": "pdf", "pages": len(reader.pages)}
        except ImportError:
            return {"success": False, "error": "未安装 PyPDF2，请运行: pip install PyPDF2"}
    
    def _read_word(self, filepath: str) -> Dict[str, Any]:
        """读取 Word"""
        try:
            import docx
            doc = docx.Document(filepath)
            text = "\n".join([para.text for para in doc.paragraphs])
            return {"success": True, "content": text, "type": "docx"}
        except ImportError:
            return {"success": False, "error": "未安装 python-docx，请运行: pip install python-docx"}
    
    def _read_excel(self, filepath: str) -> Dict[str, Any]:
        """读取 Excel"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath)
            result = {}
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                result[sheet] = data
            return {"success": True, "content": result, "type": "excel", "sheets": wb.sheetnames}
        except ImportError:
            return {"success": False, "error": "未安装 openpyxl，请运行: pip install openpyxl"}
    
    def list_session_files(self, session_id: str = "default") -> List[Dict[str, Any]]:
        """列出会话中的文件"""
        return [info for fid, info in self.file_cache.items() if info["session"] == session_id]
    
    def delete_file(self, file_id: str) -> Dict[str, Any]:
        """删除文件"""
        if file_id not in self.file_cache:
            return {"success": False, "error": "文件未找到"}
        
        try:
            info = self.file_cache[file_id]
            if os.path.exists(info["path"]):
                os.remove(info["path"])
            del self.file_cache[file_id]
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}


def initialize(plugin_manager):
    """插件初始化入口"""
    plugin = FileHandlerPlugin()
    # 注册到插件管理器
    plugin_manager.file_handler = plugin
    print("[FileHandler] Plugin loaded")
